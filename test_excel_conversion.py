"""
Excel 转 Markdown 转换测试脚本
用于验证 MarkItDown 和 openpyxl 库的 Excel 转换功能
"""

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

def create_sample_excel():
    """创建一个简单的测试 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        print("❌ openpyxl 未安装，无法创建测试文件")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试工作表"

    # 添加表头
    headers = ["姓名", "年龄", "职位", "部门"]
    ws.append(headers)

    # 添加数据行
    data = [
        ["张三", 28, "工程师", "研发部"],
        ["李四", 35, "经理", "市场部"],
        ["王五", 42, "总监", "技术部"],
        ["赵六", 31, "分析师", "数据部"],
    ]

    for row in data:
        ws.append(row)

    # 添加第二个工作表
    ws2 = wb.create_sheet("财务数据")
    ws2.append(["月份", "收入", "支出", "利润"])
    ws2.append(["1月", 100000, 80000, 20000])
    ws2.append(["2月", 120000, 90000, 30000])
    ws2.append(["3月", 150000, 100000, 50000])

    # 保存到临时文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()

    return temp_file.name


def test_markitdown_conversion(excel_path):
    """测试 MarkItDown 转换"""
    print("\n=== 测试 MarkItDown 转换 ===")
    try:
        from markitdown import MarkItDown

        print(f"✓ MarkItDown 已安装")
        print(f"正在转换: {excel_path}")

        md = MarkItDown()
        result = md.convert(excel_path)

        print(f"\n✓ MarkItDown 转换成功!")
        print(f"转换结果预览（前 500 字符）:")
        print("-" * 50)
        preview = result.text_content[:500]
        print(preview)
        if len(result.text_content) > 500:
            print("... (内容已截断)")
        print("-" * 50)

        return result.text_content
    except ImportError:
        print("⚠ MarkItDown 未安装，跳过此测试")
        return None
    except Exception as e:
        print(f"✗ MarkItDown 转换失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def test_openpyxl_conversion(excel_path):
    """测试 openpyxl 直接转换"""
    print("\n=== 测试 openpyxl 转换 ===")
    try:
        import openpyxl

        print(f"✓ openpyxl 已安装")
        print(f"正在转换: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        markdown_lines = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            markdown_lines.append(f"## {sheet_name}\n")

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if all(cell is None for cell in row):
                    continue

                cells = []
                for cell in row:
                    if cell is None:
                        cells.append("")
                    elif isinstance(cell, (int, float)):
                        cells.append(str(cell))
                    else:
                        cell_str = str(cell)
                        cell_str = cell_str.replace('|', '\\|')
                        cells.append(cell_str.strip())

                markdown_lines.append("| " + " | ".join(cells) + " |")
                if row_idx == 1:
                    markdown_lines.append("| " + " | ".join(["---"] * len(cells)) + " |")

            markdown_lines.append("")

        result = "\n".join(markdown_lines)

        print(f"\n✓ openpyxl 转换成功!")
        print(f"转换结果预览（前 500 字符）:")
        print("-" * 50)
        preview = result[:500]
        print(preview)
        if len(result) > 500:
            print("... (内容已截断)")
        print("-" * 50)

        return result
    except ImportError:
        print("⚠ openpyxl 未安装，跳过此测试")
        return None
    except Exception as e:
        print(f"✗ openpyxl 转换失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def test_excel2markdown_script(excel_path):
    """测试 excel2markdown.py 脚本"""
    print("\n=== 测试 excel2markdown.py 脚本 ===")
    try:
        from backend.scripts.excel2markdown import excel_to_markdown

        print(f"正在测试 excel_to_markdown 函数...")

        # 创建临时输出目录
        output_dir = tempfile.mkdtemp()

        # 调用转换函数
        output_path = excel_to_markdown(excel_path, output_dir)

        if output_path and os.path.exists(output_path):
            print(f"✓ 脚本执行成功!")
            print(f"输出文件: {output_path}")

            # 读取并显示内容
            with open(output_path, 'r', encoding='utf-8') as f:
                content = f.read()

            print(f"\n转换结果预览（前 500 字符）:")
            print("-" * 50)
            preview = content[:500]
            print(preview)
            if len(content) > 500:
                print("... (内容已截断)")
            print("-" * 50)

            # 清理临时文件
            os.remove(output_path)
            os.rmdir(output_dir)

            return True
        else:
            print(f"✗ 脚本执行失败，未生成输出文件")
            return False

    except Exception as e:
        print(f"✗ excel2markdown.py 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_ingest_module():
    """测试 ingest.py 中的 Excel 转换功能"""
    print("\n=== 测试 ingest.py 模块 ===")
    try:
        from backend.app.services.ingest import extract_excel_as_markdown

        print(f"✓ ingest.extract_excel_as_markdown 函数可用")

        # 创建测试文件
        excel_path = create_sample_excel()
        if not excel_path:
            print("⚠ 无法创建测试文件，跳过 ingest 模块测试")
            return False

        print(f"正在测试 extract_excel_as_markdown 函数...")

        result = extract_excel_as_markdown(excel_path)

        if result:
            print(f"✓ ingest 模块测试成功!")
            print(f"转换结果预览（前 500 字符）:")
            print("-" * 50)
            preview = result[:500]
            print(preview)
            if len(result) > 500:
                print("... (内容已截断)")
            print("-" * 50)

            # 清理临时文件
            os.remove(excel_path)

            return True
        else:
            print(f"✗ ingest 模块测试失败，未返回结果")
            os.remove(excel_path)
            return False

    except Exception as e:
        print(f"✗ ingest 模块测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    print("=" * 60)
    print("Excel 转 Markdown 转换测试")
    print("=" * 60)

    # 检查依赖
    print("\n检查依赖库...")
    deps_ok = True

    try:
        import openpyxl
        print(f"✓ openpyxl: {openpyxl.__version__}")
    except ImportError:
        print(f"✗ openpyxl: 未安装")
        deps_ok = False

    try:
        import markitdown
        print(f"✓ markitdown: 已安装")
    except ImportError:
        print(f"⚠ markitdown: 未安装（推荐安装）")

    try:
        import pandas
        print(f"✓ pandas: {pandas.__version__}")
    except ImportError:
        print(f"⚠ pandas: 未安装")

    if not deps_ok:
        print("\n❌ 错误: openpyxl 未安装，无法运行测试")
        print("请运行: pip install openpyxl")
        sys.exit(1)

    # 创建测试 Excel 文件
    print("\n创建测试 Excel 文件...")
    excel_path = create_sample_excel()
    if not excel_path:
        print("❌ 无法创建测试文件")
        sys.exit(1)

    print(f"✓ 测试文件已创建: {excel_path}")

    # 运行测试
    test_markitdown_conversion(excel_path)
    test_openpyxl_conversion(excel_path)
    test_excel2markdown_script(excel_path)
    test_ingest_module()

    # 清理
    if excel_path and os.path.exists(excel_path):
        os.remove(excel_path)

    print("\n" + "=" * 60)
    print("测试完成")
    print("=" * 60)

    print("\n📝 使用说明:")
    print("1. MarkItDown 是推荐的转换方式，支持更多格式")
    print("2. openpyxl 作为备选方案，专注于 Excel 格式")
    print("3. 如需转换 Excel 文件到知识库，只需上传 .xlsx 或 .xls 文件")
    print("4. 系统会自动调用 excel2markdown.py 进行转换")


if __name__ == "__main__":
    main()
